import os
import re
import time
from openpyxl import Workbook

pattern_pc=r'[A-Z]+(_.+?\|.+?\|turn\d+)? (G=\d+ )?\d+ \d \d\n'
pattern_pc+=r'(WISH \d \d \d \d \d \d \d\n)?'
pattern_pc+=r'(AMULET ([A-Z]+ \d+ )+ENDAMULET\n)?'
pattern_pc+=r'\d+ \d+ \d+ \d+ \d+ \d+ ?\n'
pattern_pc+=r'([A-Z]+ \d+ \d+ \d+ \d+ \d+ \d|NONE)\n'*4
pattern_pc+=r'\d( [A-Z]+)*'

class GEAR:
    def __init__(self,string):
        pattern=r'(([A-Z]+) (\d+ \d+ \d+ \d+ \d+) (\d)|NONE)'
        match=re.search(pattern,string).groups()
        self.string=match[0]
        self.type=match[1]
        self.attribute=match[2]
        if match[3]:
            self.myst=bool(int(match[3]))
        else:
            self.myst=match[3]

    def __str__(self):
        return self.string
    
    def transform(self):
        if self.myst:
            return 'M_'+self.type
        else:
            return self.type

class PC:
    def __init__(self,string):
        pattern=r'(?P<string>'
        pattern+=r'(?P<role>[A-Z]+)_(?P<name>(?P<group>.+?)\|(?P<mode>.+?)\|turn(?P<turn>\d+)) (?P<growth>G=\d+ )?(?P<card>\d+ \d \d)\n'
        pattern+=r'(?:WISH (?P<wish>\d \d \d \d \d \d \d)\n)?'
        pattern+=r'(?:AMULET (?P<amulet>(?:[A-Z]+ \d+ )+)ENDAMULET\n)?'
        pattern+=r'(?P<attribute>\d+ \d+ \d+ \d+ \d+ \d+) ?\n'
        pattern+=r'(?P<weapon>[A-Z]+ \d+ \d+ \d+ \d+ \d+ \d|NONE)\n'
        pattern+=r'(?P<hand>[A-Z]+ \d+ \d+ \d+ \d+ \d+ \d|NONE)\n'
        pattern+=r'(?P<body>[A-Z]+ \d+ \d+ \d+ \d+ \d+ \d|NONE)\n'
        pattern+=r'(?P<head>[A-Z]+ \d+ \d+ \d+ \d+ \d+ \d|NONE)\n'
        pattern+=r'(?P<aura>\d(?: [A-Z]+)*)'+r')'
        match=re.search(pattern,string)
        self.string=match['string']
        self.role=match['role']
        self.name=match['name']
        self.group=match['group']
        self.mode=match['mode']
        self.turn=match['turn']
        self.growth=match['growth']
        self.card=match['card']
        self.wish=match['wish']
        self.amulet=match['amulet']
        self.attribute=match['attribute']
        self.weapon=GEAR(match['weapon'])
        self.hand=GEAR(match['hand'])
        self.body=GEAR(match['body'])
        self.head=GEAR(match['head'])
        self.aura=match['aura']        

    def __str__(self):
        return self.string
    
    def __eq__(self,other):
        return (self.role,self.name)==(other.role,other.name)
    
    def __hash__(self):
        return hash((self.role,self.name))

def read_option():
    '''读取option.txt，返回包含信息的字典'''
    with open('option.txt',mode='r',encoding='UTF-8') as f:
        option=f.read()
    
    def read_info(info):
        '''读取信息，以给定的形式返回'''
        def read_info1(info):
            '''读取形如a=x b=y的信息，返回包含信息的字典'''
            return dict(re.findall(r'(.+)=(.+)',info))
        
        def read_info2(info):
            '''读取形如a={} b={}的信息，返回包含信息的字典'''
            match=re.findall(r'(.+)={(.+)}',info)
            dt=dict(map(lambda t:(t[0],tuple(map(lambda s:tuple(s.split(' ')),t[1].split(';')))),match))
            return dt

        def read_info3(info):
            '''读取形如{a=x b=y} {c=z d=w}的信息，返回包含信息的字典组成的元组'''
            match=re.findall(r'{(.+?)}',info,flags=re.S)
            dt=tuple(map(lambda t:dict(re.findall(r'(.+)=(.*)',t)),match))
            return dt
        
        if(not re.search(r'[{}]',info)):
            return read_info1(info)
        elif(re.search(r'={.+}',info)):
            return read_info2(info)
        elif(re.search(r'{[^{}]*=[^{}]*}',info)):
            return read_info3(info)

    pattern=r'\[(.+?)\]([^\[\]]+)'
    match=re.findall(pattern,option,flags=re.S)
    dt=dict(map(lambda t:(t[0],read_info(t[1])),match)) #返回字典，value为option中key对应的信息
    return dt

def read_pools():
    '''用户输入start和end，读取start至end的所有轮数的PC，返回key为pc、value为份额的字典'''
    dir_name=input('请输入需要统计的文件夹名（留空则默认为\'\\记录\'）：')
    if not dir_name:
        dir_name='记录'

    while(True):
        start,end=tuple(map(int,input('请输入需统计的起始轮数(start,end)：').split()))
        if(f'turn{end}.txt' in os.listdir(dir_name)):
            break
        else:
            print(f'{dir_name}\\turn{end}.txt 不存在，请重新输入！')

    dt={}
    
    for turn in range(start,end+1):
        with open(f'{dir_name}\\turn{turn}.txt',mode='r',encoding='UTF-8') as f:
            pool_all=map(lambda s:PC(s.group()),re.finditer(pattern_pc,f.read()))
            for pc in pool_all:
                if pc in dt:
                    dt[pc]+=1
                else:
                    dt[pc]=1
    
    return dt

def to_workbook(dt):
    '''将dt的内容导出至指定名称的Excel文档'''
    def mode(defender):
        '''将defender转化为mode'''
        if(defender=='0'):
            return 'ATK'
        elif(defender=='1'):
            return 'DEF'
        elif(defender=='2'):
            return 'MIX'
  
    def row_all(pc):
        return [pc.name,pc.group,pc.mode]+row_group(pc)
    
    def row_group(pc):
        row=[int(pc.turn),dt[pc],pc.role]
        row+=list(map(int,pc.attribute.split()))
        row+=[pc.weapon.transform(),pc.hand.transform(),pc.body.transform(),pc.head.transform(),pc.aura]
        row+=pc.aura.split()[1:]
        return row

    title_all=['名称','组名','攻守','轮数','份额','角色','力量','敏捷','智力','体魄','精神','意志',
                '武器','手部','身体','头部','光环','光环1','光环2','光环3','光环4']
    title_group=['轮数','份额','角色','力量','敏捷','智力','体魄','精神','意志',
                '武器','手部','身体','头部','光环','光环1','光环2','光环3','光环4']

    wb=Workbook()

    ws=wb.create_sheet('原始数据')
    ws.append(title_all)
    for group in read_option()['Group']:
        ws.temp=wb.create_sheet(group['Name']+'|'+mode(group['Defender']))
        ws.temp.append(title_group)
    for pc in dt:
        ws.append(row_all(pc))
        wb[pc.group+'|'+pc.mode].append(row_group(pc))

    xlsx_name=input('请输入导出的文件名（留空则默认为\'统计结果.xlsx\'）：')
    if not xlsx_name:
        xlsx_name='统计结果'
    wb.save(xlsx_name+'.xlsx')

def main():
    to_workbook(read_pools())

try:
    start_time=time.time()
    main()
    end_time=time.time()
    print(f'Use time: {end_time-start_time} s')
except Exception as e:
    print(e)
    input()
